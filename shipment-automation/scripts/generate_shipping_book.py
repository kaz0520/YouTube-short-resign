#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
出荷伝票 → 見本準拠Excel（配送区分ごとシート分割）生成器。
入力: JSON（配送区分ごとのグループ配列）
出力: 出荷リスト_*.xlsx

列: No / 客先番号 / 物件名・客先 / 地域 / 製品名・仕様 / 数量 / 同梱先 / 個口 / 備考
- 1出荷(No)に複数製品 → No/客先番号/物件名/地域/個口/備考 を縦結合、製品名・数量は1製品1行
- 同梱先(G列): そのグループの製品名からドロップダウン（単品でも設定）
- 要確認の製品セルは黄色
- 全セル上下中央そろえ
"""
import json
import sys
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HEADERS = ["No", "客先番号", "物件名・客先", "地域", "製品名・仕様", "数量", "同梱先", "個口", "備考"]
WIDTHS = {"A": 4, "B": 10, "C": 26.4, "D": 7.3, "E": 30, "F": 6, "G": 30, "H": 6, "I": 7}

thin = Side(style="thin")
medium = Side(style="medium")
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FONT = Font(size=14, bold=True)
HEADER_FONT = Font(bold=True)

# 列ごとの水平配置 / 折り返し
HALIGN = {1: "center", 2: "center", 3: "left", 4: "center", 5: "left",
          6: "center", 7: "left", 8: "center", 9: "left"}
WRAP = {3, 5}


def al(col, header=False):
    h = "center" if header else HALIGN[col]
    return Alignment(horizontal=h, vertical="center", wrap_text=(col in WRAP))


def border(bottom=thin):
    return Border(left=thin, right=thin, top=thin, bottom=bottom)


def build_sheet(ws, title, groups):
    # タイトル
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    c = ws.cell(1, 1, title)
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ヘッダー
    for col, name in enumerate(HEADERS, 1):
        cell = ws.cell(2, col, name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = al(col, header=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=medium)

    r = 3
    for g in groups:
        prods = g["products"] if g["products"] else [{"name": "", "qty": ""}]
        n = len(prods)
        top, bot = r, r + n - 1
        names = [p["name"] for p in prods if p.get("name")]
        # プルダウン（同梱先）: グループ内製品から
        dv = None
        if names:
            formula = '"' + ",".join(x.replace(",", " ").replace('"', "") for x in names)[:250] + '"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(dv)

        for i, p in enumerate(prods):
            rr = r + i
            is_group_last = (rr == bot)
            bb = medium if is_group_last else thin
            # 製品名(E) 数量(F) 同梱先(G)
            ename = p["name"]
            if p.get("note"):
                ename = f"{ename}\n{p['note']}" if ename else p["note"]
            ce = ws.cell(rr, 5, ename)
            ce.alignment = al(5)
            ce.border = border(bb)
            if p.get("warn"):
                ce.fill = WARN_FILL
            qv = p.get("qty", "")
            cf = ws.cell(rr, 6, qv if qv != "" else None)
            cf.alignment = al(6)
            cf.border = border(bb)
            cg = ws.cell(rr, 7, "")
            cg.alignment = al(7)
            cg.border = border(bb)
            if dv:
                dv.add(cg)
        # 結合列: No/客先番号/物件名/地域/個口/備考
        merged = {1: g.get("no"), 2: g.get("kokyaku"), 3: g.get("bukken"),
                  4: g.get("chiiki"), 8: g.get("koguchi"), 9: g.get("bikou")}
        for col, val in merged.items():
            if n > 1:
                ws.merge_cells(start_row=top, start_column=col, end_row=bot, end_column=col)
            cell = ws.cell(top, col, val if val not in (None, "") else None)
            cell.alignment = al(col)
            if col == 3 and g.get("bukken_warn"):
                cell.fill = WARN_FILL
            # 罫線: 結合セルは範囲全体に付与
            for rr in range(top, bot + 1):
                bb = medium if rr == bot else thin
                ws.cell(rr, col).border = border(bb)
        r = bot + 1

    for col_letter, w in WIDTHS.items():
        ws.column_dimensions[col_letter].width = w
    # 行高さ自動＝openpyxlでは明示不可のため、折り返しはExcel側で調整


def main(data_path, out_path):
    data = json.load(open(data_path, encoding="utf-8"))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet in data["sheets"]:
        ws = wb.create_sheet(sheet["name"])
        build_sheet(ws, sheet["title"], sheet["groups"])
    # 要確認シート
    if data.get("review"):
        ws = wb.create_sheet("要確認")
        ws.append(["シート", "No", "客先/物件", "対象", "内容"])
        for col in range(1, 6):
            cc = ws.cell(1, col)
            cc.fill = HEADER_FILL
            cc.font = HEADER_FONT
            cc.alignment = Alignment(horizontal="center", vertical="center")
        for row in data["review"]:
            ws.append(row)
        for w, c in zip([12, 6, 22, 22, 40], "ABCDE"):
            ws.column_dimensions[c].width = w
    wb.save(out_path)
    print(f"SUCCESS:{out_path}  sheets={[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
